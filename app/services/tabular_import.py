import csv
import io
from typing import Any

from openpyxl import load_workbook
import xlrd


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _looks_like_header(row: list[Any]) -> bool:
    vals = [_normalize_header(x) for x in row]
    if len([v for v in vals if v]) < 3:
        return False
    keywords = [
        "артикул",
        "sku",
        "код",
        "наименование",
        "товар",
        "бренд",
        "производитель",
        "цена",
        "группа",
    ]
    hits = sum(1 for v in vals if any(k in v for k in keywords))
    return hits >= 2


def parse_table_bytes(filename: str, payload: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        text = payload.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or "").strip(): v for k, v in row.items()} for row in reader]
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header_idx = 0
        for i, row in enumerate(rows[:40]):
            if _looks_like_header(list(row)):
                header_idx = i
                break
        headers = [_normalize_header(x) for x in rows[header_idx]]
        result: list[dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            item: dict[str, Any] = {}
            for idx, val in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
                item[key] = val
            if any(v is not None and str(v).strip() for v in item.values()):
                result.append(item)
        return result
    if lower.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=payload)
        ws = wb.sheet_by_index(0)
        if ws.nrows == 0:
            return []
        header_idx = 0
        for r in range(min(ws.nrows, 40)):
            row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
            if _looks_like_header(row_vals):
                header_idx = r
                break
        headers = [_normalize_header(ws.cell_value(header_idx, c)) for c in range(ws.ncols)]
        result: list[dict[str, Any]] = []
        for r in range(header_idx + 1, ws.nrows):
            item: dict[str, Any] = {}
            for c in range(ws.ncols):
                key = headers[c] if c < len(headers) else f"col_{c + 1}"
                item[key] = ws.cell_value(r, c)
            if any(v is not None and str(v).strip() for v in item.values()):
                result.append(item)
        return result
    raise ValueError("Unsupported import format. Use CSV/XLSX/XLS.")


def to_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "да"}:
        return True
    if text in {"0", "false", "no", "n", "нет"}:
        return False
    return default


def to_float(value: Any, default: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return default
